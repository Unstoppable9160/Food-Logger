import os
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

os.system("cls")

name = str(input("Please enter the date without any special characters (eg: 3rd July): "))
dbname = str(f"Food Log {name}")
i = 2
text = """
		     	Food Log			

Log the amount of calories you consumed throughout the day,
whenever you eat just type in the calories and dont think about it!

Format: food name|calories

use the '|' to split the 2 values

Type quit when you are done with the day...
"""
heading = Font(name='Arial', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False)


def saveWorkbook(name_: str, workbook: openpyxl.Workbook):
    workbook.save(f"{name_}.xlsx")
    print(f"Workbook '{name_}' created")
    return True

def write(cell:str, text):
    ws[cell] = text


wb = Workbook()
ws = wb.active


write('A1', "Time")
write('B1', "Food Name")
write('C1', "Calories")
a1 = ws['A1']
b1 = ws['B1']
c1 = ws['C1']

a1.font = heading
b1.font = heading
c1.font = heading

print(text)
while True:
    userinput = input("Please enter the name of the food and the amount of calories as integers split by '|': ")
    if userinput != "quit":
        foodName, calories = userinput.split('|')
        now = datetime.datetime.now()
        current_time = now.strftime("%H:%M:%S")
        write(f'A{i}', current_time)
        write(f'B{i}', foodName)
        write(f'C{i}', calories)
        i = i + 1
    if userinput == "quit":
        break

saveWorkbook(dbname, wb)
